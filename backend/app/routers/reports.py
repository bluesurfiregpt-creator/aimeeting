"""
v23 + v24.3 #2 — 报表导出 (Excel only).

设计哲学:
- 政务用户拿到报表后会自己改表头 / 着色 / 挑数据 → Excel 优于 PDF (灵活)
- 后端用 openpyxl 生成 (已经在依赖里 — KB 文档解析也用)
- 表头加粗 + 冻结首行,基本可读
- 文件名带时间戳防覆盖

输出(智慧住建文档 §4.5 视图 4 报表体系 — 日清/周查/月结):
  /api/reports/daily-summary?date=YYYY-MM-DD       v24.3 日清 — 当日 Task 全表 + 当日新建 / 完成 / 逾期 计数
  /api/reports/weekly-summary?week_start=YYYY-MM-DD v24.3 周查 — 周内 7 天日维 + 周末状态总分布 + Agent 工作量
  /api/reports/monthly-evaluation?period=YYYY-MM    v23   月结 — 全员 4 维评价
  /api/reports/status-distribution?days=30          v23   末 N 天状态分布
"""

from __future__ import annotations

import io
import logging
from datetime import date, datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from ..auth import AuthContext, get_current_auth, require_leader_or_admin
from ..db import get_session
from ..models import Agent, Task, TaskEvaluation, User, WorkspaceMembership

router = APIRouter(prefix="/api/reports", tags=["reports"])
logger = logging.getLogger(__name__)

# Excel content-type & 中文文件名 (RFC 5987)
_XLSX_CT = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    """openpyxl Workbook → StreamingResponse (utf-8 文件名).

    Starlette 把 headers 强制 encode 成 latin-1,所以裸中文 filename 会爆
    UnicodeEncodeError.正确做法:
      - `filename=` 段必须 ASCII 安全(给老浏览器看)
      - `filename*=UTF-8''<percent-encoded>` 段给 modern 浏览器(RFC 5987),
        decode 后才是真正的中文文件名

    实测 Chrome / Edge / Firefox / Safari 都优先取 filename*= 后呈现.
    """
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote

    encoded = quote(filename)
    # ASCII fallback — 把所有非 ASCII 字符替换成下划线,保证 latin-1 可编码
    ascii_fallback = filename.encode("ascii", errors="replace").decode("ascii").replace("?", "_")
    return StreamingResponse(
        buf,
        media_type=_XLSX_CT,
        headers={
            "Content-Disposition": (
                f"attachment; filename=\"{ascii_fallback}\"; "
                f"filename*=UTF-8''{encoded}"
            ),
        },
    )


# ---- 月度评价导出 ----------------------------------------------------------


_HEADER_FILL = PatternFill(start_color="FF1F2430", end_color="FF1F2430", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_CENTER = Alignment(horizontal="center", vertical="center")


def _autosize_columns(ws, max_width: int = 40) -> None:
    """简易列宽自适应 — openpyxl 没有 auto-fit,用最长字符数估算."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=0,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)


@router.get("/monthly-evaluation")
async def monthly_evaluation_xlsx(
    period: Optional[str] = Query(
        None, regex=r"^\d{4}-\d{2}$", description="YYYY-MM,默认本月"
    ),
    session: AsyncSession = Depends(get_session),
    auth: AuthContext = Depends(get_current_auth),
):
    """
    v23: 月度 4 维评价导出 (Excel).
    leader/admin only — 个人和 expert 看自己的就够,不需要全员表.
    """
    await require_leader_or_admin(session, auth)
    p = period or datetime.now(timezone.utc).strftime("%Y-%m")

    rows = (
        await session.execute(
            select(TaskEvaluation, User)
            .join(User, User.id == TaskEvaluation.assignee_user_id)
            .where(
                TaskEvaluation.workspace_id == auth.workspace.id,
                TaskEvaluation.period == p,
            )
            .order_by(
                (
                    TaskEvaluation.completion_rate * 0.3
                    + TaskEvaluation.on_time_rate * 0.3
                    + TaskEvaluation.quality_score * 0.2
                    + TaskEvaluation.collaboration_score * 0.2
                ).desc()
            )
        )
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"{p} 月度评价"

    # 顶部 meta(merge cells)
    ws.cell(row=1, column=1, value=f"工作空间:{auth.workspace.name}").font = Font(bold=True)
    ws.cell(row=2, column=1, value=f"周期:{p}")
    ws.cell(row=3, column=1, value=f"导出时间:{datetime.now(timezone.utc).astimezone().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(
        row=4,
        column=1,
        value="综合分 = 完成率 ×30% + 及时率 ×30% + 质量 ×20% + 协作 ×20%",
    ).font = Font(italic=True, color="888888")

    # 表头(第 6 行)
    headers = [
        "排名", "姓名", "完成率", "及时率", "质量分", "协作分",
        "综合分", "本月分配", "本月完成", "本月超期",
    ]
    header_row = 6
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER

    # 数据行
    for i, (e, u) in enumerate(rows, start=1):
        composite = round(
            e.completion_rate * 0.3
            + e.on_time_rate * 0.3
            + e.quality_score * 0.2
            + e.collaboration_score * 0.2,
            3,
        )
        ws.cell(row=header_row + i, column=1, value=i)
        ws.cell(row=header_row + i, column=2, value=u.name)
        ws.cell(row=header_row + i, column=3, value=round(e.completion_rate, 3))
        ws.cell(row=header_row + i, column=4, value=round(e.on_time_rate, 3))
        ws.cell(row=header_row + i, column=5, value=round(e.quality_score, 3))
        ws.cell(row=header_row + i, column=6, value=round(e.collaboration_score, 3))
        ws.cell(row=header_row + i, column=7, value=composite)
        ws.cell(row=header_row + i, column=8, value=e.total_assigned or 0)
        ws.cell(row=header_row + i, column=9, value=e.total_done or 0)
        ws.cell(row=header_row + i, column=10, value=e.total_overdue or 0)

    # 数值列百分比格式
    for col in range(3, 8):
        for r in range(header_row + 1, header_row + 1 + len(rows)):
            cell = ws.cell(row=r, column=col)
            cell.number_format = "0.0%"

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _autosize_columns(ws)

    fname = f"月度评价_{auth.workspace.name}_{p}_{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return _xlsx_response(wb, fname)


# ---- 状态分布趋势导出 -----------------------------------------------------


@router.get("/status-distribution")
async def status_distribution_xlsx(
    days: int = Query(30, ge=7, le=180),
    session: AsyncSession = Depends(get_session),
    auth: AuthContext = Depends(get_current_auth),
):
    """
    v23: 末 N 天状态分布趋势导出 (Excel,leader/admin only).

    每行一天,各状态计数 + 当天创建数 + 当天完成数.
    便于看「这 X 天积压有没有改善」.
    """
    await require_leader_or_admin(session, auth)
    now = datetime.now(timezone.utc)
    start = now - timedelta(days=days)

    # 按 (date, status) 聚合 created
    created_rows = (
        await session.execute(
            select(
                func.date(Task.created_at).label("d"),
                Task.status,
                func.count(Task.id),
            )
            .where(
                Task.workspace_id == auth.workspace.id,
                Task.created_at >= start,
            )
            .group_by("d", Task.status)
            .order_by("d")
        )
    ).all()
    # 每天 done 计数(updated_at 为终结时间)
    done_rows = (
        await session.execute(
            select(
                func.date(Task.updated_at).label("d"),
                func.count(Task.id),
            )
            .where(
                Task.workspace_id == auth.workspace.id,
                Task.status == "done",
                Task.updated_at >= start,
            )
            .group_by("d")
            .order_by("d")
        )
    ).all()

    # 索引化
    created_by_day_status: dict[str, dict[str, int]] = {}
    for d, s, n in created_rows:
        created_by_day_status.setdefault(str(d), {})[s] = int(n)
    done_by_day = {str(d): int(n) for d, n in done_rows}

    wb = Workbook()
    ws = wb.active
    ws.title = f"末 {days} 天状态分布"

    ws.cell(row=1, column=1, value=f"工作空间:{auth.workspace.name}").font = Font(bold=True)
    ws.cell(row=2, column=1, value=f"区间:近 {days} 天")
    ws.cell(row=3, column=1, value=f"导出时间:{now.astimezone().strftime('%Y-%m-%d %H:%M')}")

    statuses = ["open", "dispatched", "accepted", "in_progress", "submitted", "done", "archived", "cancelled"]
    cn_status = {
        "open": "未派发", "dispatched": "待签收", "accepted": "已签收",
        "in_progress": "办理中", "submitted": "待审核", "done": "已完成",
        "archived": "已归档", "cancelled": "已取消",
    }

    headers = ["日期", *[cn_status[s] for s in statuses], "当日新建总数", "当日完成数"]
    header_row = 5
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER

    # 把 N 天每天填齐
    for i in range(days, -1, -1):
        d_iso = (now - timedelta(days=i)).date().isoformat()
        row_idx = header_row + (days - i) + 1
        ws.cell(row=row_idx, column=1, value=d_iso)
        per_status = created_by_day_status.get(d_iso, {})
        total_created = 0
        for j, s in enumerate(statuses, start=2):
            n = per_status.get(s, 0)
            total_created += n
            ws.cell(row=row_idx, column=j, value=n)
        ws.cell(row=row_idx, column=2 + len(statuses), value=total_created)
        ws.cell(row=row_idx, column=3 + len(statuses), value=done_by_day.get(d_iso, 0))

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)
    _autosize_columns(ws)

    fname = f"状态分布_{auth.workspace.name}_近{days}天_{now.strftime('%Y%m%d-%H%M')}.xlsx"
    return _xlsx_response(wb, fname)


# ---- v24.3 #2: 日清(daily summary)+ 周查(weekly summary)---------------

_STATUS_CN = {
    "open": "未派发", "dispatched": "待签收", "accepted": "已签收",
    "in_progress": "办理中", "submitted": "待审核", "done": "已完成",
    "archived": "已归档", "cancelled": "已取消",
}
_SOURCE_CN = {
    "meeting": "会议", "manual": "手工", "leader_directive": "领导指令",
    "upper_doc": "上级文件", "cron": "定期巡检",
    "alert": "异常预警", "report": "问题上报",
}


def _parse_iso_date(s: Optional[str]) -> date:
    if not s:
        return datetime.now(timezone.utc).date()
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, f"date 格式应为 YYYY-MM-DD,实际 {s}")


@router.get("/daily-summary")
async def daily_summary_xlsx(
    date: Optional[str] = Query(None, description="YYYY-MM-DD,默认今日"),
    session: AsyncSession = Depends(get_session),
    auth: AuthContext = Depends(get_current_auth),
):
    """
    v24.3 #2 日清(智慧住建文档 §4.5 视图 4):

    - Sheet 1「当日 Task 全表」:当日新建的所有 Task 一行一份(含 ID / 标题 /
      状态 / 来源 / assignee / 创建时间 / 截止时间)
    - Sheet 2「当日汇总」:新建 / 完成 / 逾期 / 状态分布 计数

    leader/admin only.
    """
    await require_leader_or_admin(session, auth)
    d = _parse_iso_date(date)
    day_start = datetime.combine(d, datetime.min.time(), tzinfo=timezone.utc)
    day_end = day_start + timedelta(days=1)

    # ---- Sheet 1:当日新建 Task 全表
    rows = (
        await session.execute(
            select(Task)
            .where(
                Task.workspace_id == auth.workspace.id,
                Task.created_at >= day_start,
                Task.created_at < day_end,
            )
            .order_by(Task.created_at)
        )
    ).scalars().all()
    user_ids = {t.assignee_user_id for t in rows if t.assignee_user_id}
    name_by_uid: dict = {}
    if user_ids:
        urows = (
            await session.execute(select(User).where(User.id.in_(user_ids)))
        ).scalars().all()
        name_by_uid = {u.id: u.name for u in urows}

    wb = Workbook()
    ws = wb.active
    ws.title = f"{d.isoformat()} 任务"

    ws.cell(row=1, column=1, value=f"工作空间:{auth.workspace.name}").font = Font(bold=True)
    ws.cell(row=2, column=1, value=f"日期:{d.isoformat()}")
    ws.cell(row=3, column=1, value=f"导出时间:{datetime.now(timezone.utc).astimezone().strftime('%Y-%m-%d %H:%M')}")

    headers = ["任务 ID", "标题", "状态", "来源", "主责", "创建时间", "截止时间"]
    header_row = 5
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER

    for i, t in enumerate(rows, start=1):
        title = (t.title or t.content[:40]).replace("\n", " ")
        ws.cell(row=header_row + i, column=1, value=str(t.id)[:8])
        ws.cell(row=header_row + i, column=2, value=title[:80])
        ws.cell(row=header_row + i, column=3, value=_STATUS_CN.get(t.status, t.status))
        ws.cell(row=header_row + i, column=4, value=_SOURCE_CN.get(t.source_type, t.source_type))
        ws.cell(row=header_row + i, column=5, value=name_by_uid.get(t.assignee_user_id, "-") if t.assignee_user_id else "-")
        ws.cell(row=header_row + i, column=6,
                value=t.created_at.astimezone().strftime("%H:%M"))
        ws.cell(row=header_row + i, column=7,
                value=t.due_at.astimezone().strftime("%Y-%m-%d") if t.due_at else "-")

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _autosize_columns(ws)

    # ---- Sheet 2:当日汇总
    ws2 = wb.create_sheet(title="当日汇总")
    # 当日完成数(updated_at 在该日 + status=done)
    done_today = (
        await session.execute(
            select(func.count(Task.id)).where(
                Task.workspace_id == auth.workspace.id,
                Task.status == "done",
                Task.updated_at >= day_start,
                Task.updated_at < day_end,
            )
        )
    ).scalar() or 0
    # 当日时刻 active task 中已逾期数
    overdue_now = (
        await session.execute(
            select(func.count(Task.id)).where(
                Task.workspace_id == auth.workspace.id,
                Task.due_at.is_not(None),
                Task.due_at < day_end,
                Task.status.notin_(("done", "archived", "cancelled")),
            )
        )
    ).scalar() or 0
    # 当日各状态新建数
    by_status_rows = (
        await session.execute(
            select(Task.status, func.count(Task.id))
            .where(
                Task.workspace_id == auth.workspace.id,
                Task.created_at >= day_start,
                Task.created_at < day_end,
            )
            .group_by(Task.status)
        )
    ).all()
    summary_pairs: list[tuple[str, int | str]] = [
        ("当日新建总数", len(rows)),
        ("当日完成数", int(done_today)),
        ("截至当日活跃中已逾期", int(overdue_now)),
        ("", ""),
        ("--- 当日新建按状态 ---", ""),
    ]
    for s, c in by_status_rows:
        summary_pairs.append((_STATUS_CN.get(s, s), int(c)))
    for i, (k, v) in enumerate(summary_pairs, start=1):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True) if k.startswith("---") or k.endswith("总数") else Font()
        ws2.cell(row=i, column=2, value=v)
    _autosize_columns(ws2)

    fname = f"日清_{auth.workspace.name}_{d.isoformat()}_{datetime.now().strftime('%H%M')}.xlsx"
    return _xlsx_response(wb, fname)


@router.get("/weekly-summary")
async def weekly_summary_xlsx(
    week_start: Optional[str] = Query(None, description="YYYY-MM-DD 周一,默认本周一"),
    session: AsyncSession = Depends(get_session),
    auth: AuthContext = Depends(get_current_auth),
):
    """
    v24.3 #2 周查(智慧住建文档 §4.5 视图 4):

    - Sheet 1「周内 7 天日维」:每天一行 — 新建数 / 完成数 / 当日逾期
    - Sheet 2「Agent 工作量」:本周新建 Task 按 assignee.bound_agent 分组
    - Sheet 3「Top 10 待办」:本周末仍 active 的高优先级 task 前 10(按
      due_at 升序;无 due_at 排末尾)

    leader/admin only.
    """
    await require_leader_or_admin(session, auth)
    if week_start:
        ws_date = _parse_iso_date(week_start)
    else:
        today = datetime.now(timezone.utc).date()
        ws_date = today - timedelta(days=today.weekday())  # 本周一
    week_start_dt = datetime.combine(ws_date, datetime.min.time(), tzinfo=timezone.utc)
    week_end_dt = week_start_dt + timedelta(days=7)

    wb = Workbook()

    # ---- Sheet 1:周内 7 天日维
    ws1 = wb.active
    ws1.title = "周内 7 天"
    ws1.cell(row=1, column=1, value=f"工作空间:{auth.workspace.name}").font = Font(bold=True)
    ws1.cell(row=2, column=1, value=f"周次:{ws_date.isoformat()} 至 {(ws_date + timedelta(days=6)).isoformat()}")
    ws1.cell(row=3, column=1, value=f"导出时间:{datetime.now(timezone.utc).astimezone().strftime('%Y-%m-%d %H:%M')}")

    headers = ["日期", "星期", "新建", "完成", "当日活跃逾期"]
    header_row = 5
    for col, h in enumerate(headers, start=1):
        c = ws1.cell(row=header_row, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
    weekday_cn = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    for i in range(7):
        d = ws_date + timedelta(days=i)
        d_start = datetime.combine(d, datetime.min.time(), tzinfo=timezone.utc)
        d_end = d_start + timedelta(days=1)
        created = (
            await session.execute(
                select(func.count(Task.id)).where(
                    Task.workspace_id == auth.workspace.id,
                    Task.created_at >= d_start,
                    Task.created_at < d_end,
                )
            )
        ).scalar() or 0
        done = (
            await session.execute(
                select(func.count(Task.id)).where(
                    Task.workspace_id == auth.workspace.id,
                    Task.status == "done",
                    Task.updated_at >= d_start,
                    Task.updated_at < d_end,
                )
            )
        ).scalar() or 0
        overdue = (
            await session.execute(
                select(func.count(Task.id)).where(
                    Task.workspace_id == auth.workspace.id,
                    Task.due_at.is_not(None),
                    Task.due_at < d_end,
                    Task.status.notin_(("done", "archived", "cancelled")),
                )
            )
        ).scalar() or 0
        ws1.cell(row=header_row + 1 + i, column=1, value=d.isoformat())
        ws1.cell(row=header_row + 1 + i, column=2, value=weekday_cn[i])
        ws1.cell(row=header_row + 1 + i, column=3, value=int(created))
        ws1.cell(row=header_row + 1 + i, column=4, value=int(done))
        ws1.cell(row=header_row + 1 + i, column=5, value=int(overdue))
    ws1.freeze_panes = ws1.cell(row=header_row + 1, column=1)
    _autosize_columns(ws1)

    # ---- Sheet 2:Agent 工作量(本周新建按 bound_agent)
    ws2 = wb.create_sheet(title="Agent 工作量")
    agent_rows = (
        await session.execute(
            select(Agent.name, func.count(Task.id))
            .select_from(Task)
            .join(WorkspaceMembership, WorkspaceMembership.user_id == Task.assignee_user_id)
            .join(Agent, Agent.id == WorkspaceMembership.bound_agent_id)
            .where(
                Task.workspace_id == auth.workspace.id,
                Task.created_at >= week_start_dt,
                Task.created_at < week_end_dt,
                WorkspaceMembership.workspace_id == auth.workspace.id,
                WorkspaceMembership.bound_agent_id.is_not(None),
            )
            .group_by(Agent.name)
            .order_by(func.count(Task.id).desc())
        )
    ).all()
    ws2.cell(row=1, column=1, value="AI 专家").font = _HEADER_FONT
    ws2.cell(row=1, column=1).fill = _HEADER_FILL
    ws2.cell(row=1, column=2, value="本周新建数").font = _HEADER_FONT
    ws2.cell(row=1, column=2).fill = _HEADER_FILL
    for i, (name, c) in enumerate(agent_rows, start=2):
        ws2.cell(row=i, column=1, value=name)
        ws2.cell(row=i, column=2, value=int(c))
    if not agent_rows:
        ws2.cell(row=2, column=1, value="(本周尚无任务关联到任何 AI 专家)")
    _autosize_columns(ws2)

    # ---- Sheet 3:Top 10 待办(周末仍 active,按 due_at 升序)
    ws3 = wb.create_sheet(title="Top 10 待办")
    top_rows = (
        await session.execute(
            select(Task)
            .where(
                Task.workspace_id == auth.workspace.id,
                Task.status.in_(("open", "dispatched", "accepted", "in_progress", "submitted")),
            )
            .order_by(Task.due_at.asc().nullslast(), Task.created_at.desc())
            .limit(10)
        )
    ).scalars().all()
    user_ids = {t.assignee_user_id for t in top_rows if t.assignee_user_id}
    name_by_uid: dict = {}
    if user_ids:
        urows = (
            await session.execute(select(User).where(User.id.in_(user_ids)))
        ).scalars().all()
        name_by_uid = {u.id: u.name for u in urows}
    headers3 = ["任务", "状态", "主责", "截止时间", "已逾期"]
    for col, h in enumerate(headers3, start=1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
    now_check = datetime.now(timezone.utc)
    for i, t in enumerate(top_rows, start=2):
        title = (t.title or t.content[:50]).replace("\n", " ")
        ws3.cell(row=i, column=1, value=title[:80])
        ws3.cell(row=i, column=2, value=_STATUS_CN.get(t.status, t.status))
        ws3.cell(row=i, column=3, value=name_by_uid.get(t.assignee_user_id, "-") if t.assignee_user_id else "-")
        ws3.cell(row=i, column=4,
                 value=t.due_at.astimezone().strftime("%Y-%m-%d") if t.due_at else "-")
        ws3.cell(row=i, column=5,
                 value="是" if (t.due_at and t.due_at < now_check) else "否")
    _autosize_columns(ws3)

    fname = f"周查_{auth.workspace.name}_{ws_date.isoformat()}_{datetime.now().strftime('%H%M')}.xlsx"
    return _xlsx_response(wb, fname)
